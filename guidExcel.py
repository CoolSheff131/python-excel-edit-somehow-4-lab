import sys

import openpyxl
import xlrd
from PyQt5.QtCore import Qt

from design import Ui_Dialog

from PyQt5.QtWidgets import (
    QApplication, QDialog, QMainWindow, QMessageBox, QTableWidgetItem
)
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.uic import loadUi
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import pandas as pd  # pip install pandas
import numpy as np
from xlrd import open_workbook
from openpyxl import load_workbook, Workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border

class Window(QMainWindow, Ui_Dialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.wbHappyData = None
        self.columns = None
        self.setupUi(self)
        self.selectedCell = None

        self.loadRazdachaBtn.clicked.connect(self.loadRazdacha)
        self.loadHappyBtn.clicked.connect(self.loadHappyData)
        self.saveRazdachaBtn.clicked.connect(self.saveRazdacha)

        self.makeGreenButton.clicked.connect(self.makeCellGreen)
        self.makeYellowButton.clicked.connect(self.makeCellYellow)
        self.makeRedButton.clicked.connect(self.makeCellRed)


        self.path1 = './self.path1.xlsx'

        self.path2 = './2.xlsx'

    def makeCellGreen(self):
        if self.selectedCell == None:
            return
        self.selectedCell.setBackground(Qt.green);

    def makeCellYellow(self):
        if self.selectedCell == None:
            return
        self.selectedCell.setBackground(Qt.yellow);

    def makeCellRed(self):
        if self.selectedCell == None:
            return
        self.selectedCell.setBackground(Qt.red);

    ### По нажатию на ячейку в таблице раздачи, взять все строки в которых есть имя покупателя из happy.
    def cellClicked(self, row, column):
        table = self.tabWidget.currentWidget()
        print(row, column)
        self.selectedCell = table.item(row,column)
        return

        while (self.table_3.rowCount() > 0):
            self.table_3.removeRow(0)

        rowCount = self.table_2.rowCount()
        columnCount = self.table_2.columnCount()
        self.table_3.setRowCount(rowCount)
        self.table_3.setColumnCount(columnCount)

        clickedRowText = table.item(row, 1).text()
        print(clickedRowText)
        self.label_3.setText(clickedRowText)
        lastRowInsertedIndex = 0
        for rowIndex in range(rowCount):
            isInsertRow = False
            print(rowIndex)
            for columnIndex in range(columnCount):
                print(rowIndex, columnIndex)
                print(columnCount)
                cell = self.table_2.item(rowIndex, columnIndex)
                print(cell)
                if cell != None:
                    cellText = cell.text()
                    print(cellText)

                    if cellText != None and cellText == clickedRowText:
                        isInsertRow = True
                        break
            print(isInsertRow)

            if isInsertRow:
                for columnIndex in range(columnCount):
                    cell = self.table_2.item(rowIndex, columnIndex)
                    if cell != None:
                        cellText = cell.text()
                        tableItem = QTableWidgetItem(cellText)
                        self.table_3.setItem(lastRowInsertedIndex, columnIndex, tableItem)
                lastRowInsertedIndex += 1

        self.table_3.setColumnWidth(2, 300)

    def saveRazdacha(self):
        filename = '2self.path1.xlsx'
        print('Save razdacha')
        wb = Workbook()
        wb.remove(wb.active)
        for tabIndex in range (self.tabWidget.count()):
            print(self.tabWidget.tabText(tabIndex))
            sheet = wb.create_sheet(str(self.tabWidget.tabText(tabIndex)))

            table = self.tabWidget.widget(tabIndex)


            for row in range(table.rowCount()):
                for column in range(table.columnCount()):
                    text = table.item(row, column).text()

                    try:
                        text = str(table.item(row, column).text())
                        sheet.cell(row+1, column+1).value = text

                        background = table.item(row, column).background()
                        redFill = PatternFill(start_color='FFFFFFFF',
                                              end_color='FFFFFFFF',
                                              fill_type='solid')

                        if (background == Qt.yellow):
                            redFill = PatternFill(start_color='FFFFFF00',
                                                  end_color='FFFFFF00',
                                                  fill_type='solid')
                        elif (background == Qt.red):
                            redFill = PatternFill(start_color='FFFF0000',
                                                  end_color='FFFF0000',
                                                  fill_type='solid')
                        elif (background == Qt.green):
                            redFill = PatternFill(start_color='FF00ff00',
                                                  end_color='FF00ff00',
                                                  fill_type='solid')

                        sheet.cell(row+1, column+1).fill = redFill
                    except AttributeError:
                        pass

        wb.save(filename)

    ### Загрузить razdacha.xlsx.
    def loadRazdacha(self):
        while (self.table_3.rowCount() > 0):
            self.table_3.removeRow(0)


        sheetName = 'дост26.12 (2)'

        wb_obj = openpyxl.load_workbook(self.path1, data_only=True)

        self.tabWidget.clear()
        for sheetItemName in wb_obj.sheetnames:
            tab = QTabWidget()
            table = QTableWidget()
            self.tabWidget.addTab(table, sheetItemName)
            sheet_obj = wb_obj[sheetItemName]
            m_row = sheet_obj.max_row
            m_column = sheet_obj.max_row
            table.cellClicked.connect(self.cellClicked)

            table.setRowCount(m_row)
            table.setColumnCount(m_column)
            for i in range(1, m_row + 1):
                for j in range(1, m_column + 1):
                    cell_obj = sheet_obj.cell(row=i, column=j)
                    tableItem = QTableWidgetItem(str(cell_obj.value or ''))
                    table.setItem(i - 1, j - 1, tableItem)
                    table
                    colorHex = cell_obj.fill.start_color.index
                    if (colorHex == 'FFFFFF00'):
                        tableItem.setBackground(Qt.yellow);
                    elif (colorHex == 7):
                        tableItem.setBackground(Qt.red);
                    elif (colorHex == 'FF92D050'):
                        tableItem.setBackground(Qt.green);
            table.setColumnWidth(2, 300)

            # tab.children().append(table)
        self.wb1 = wb_obj



    ### Загруить happy.xlsx. Все страницы в таблицу
    def loadHappyData(self):
        wb_obj = openpyxl.load_workbook(self.path2)
        self.wbHappyData = wb_obj
        rowCount = 0
        columnMaximum = 0
        for sheet in wb_obj.worksheets:
            rowCount += sheet.max_row
            if sheet.max_column > columnMaximum:
                columnMaximum = sheet.max_column

        self.table_2.setRowCount(rowCount)
        self.table_2.setColumnCount(columnMaximum)
        rowIndex = 0
        for sheet in wb_obj.worksheets:
            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    cell_obj = sheet.cell(row=i, column=j)
                    tableItem = QTableWidgetItem(str(cell_obj.value or ''))

                    self.table_2.setItem(rowIndex, j - 1, tableItem)

                    colorHex = cell_obj.fill.start_color.index
                    if (colorHex == 'FFFFFF00'):
                        tableItem.setBackground(Qt.yellow);
                    elif (colorHex == 7):
                        tableItem.setBackground(Qt.red);
                    elif (colorHex == 'FF92D050'):
                        tableItem.setBackground(Qt.green);
                rowIndex += 1
        self.table_2.setColumnWidth(2, 300)




if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Window()
    win.show()
    sys.exit(app.exec())
